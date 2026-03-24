"""
test_output.py — Unit tests for output_csv.py and output_xlsx.py.

All tests use in-memory fake ScoredJob objects — no real files or API calls.
Run with:
    pytest tests/test_output.py -v
"""

import csv
import os
import tempfile
from pathlib import Path

import pytest

from src.score_jobs import ScoredJob


# ── Fixtures ───────────────────────────────────────────────────────────────────

def _make_job(
    title="Data Scientist",
    company="Flipkart",
    fit_score=0.82,
    matched=None,
    missing=None,
    salary_min=None,
    salary_max=None,
    salary_note=None,
    url="https://example.com/job/1",
    llm_summary="Strong Python and ML fit.",
    source="adzuna",
):
    return ScoredJob(
        id="abc123",
        title=title,
        company=company,
        location="Bengaluru, India",
        description="A data scientist role.",
        url=url,
        salary_min=salary_min,
        salary_max=salary_max,
        salary_note=salary_note,
        posted_date="2025-01-15",
        source=source,
        fit_score=fit_score,
        semantic_score=fit_score,
        skill_score=fit_score,
        matched_skills=matched or ["Python", "SQL", "Machine Learning"],
        missing_skills=missing or ["Spark", "Kafka"],
        bonus_skills=["Tableau"],
        llm_summary=llm_summary,
    )


def _make_jobs(n=5, base_score=0.80):
    """Create n fake jobs with descending fit scores."""
    return [
        _make_job(
            title=f"Role {i}",
            company=f"Company {i}",
            fit_score=round(base_score - i * 0.05, 3),
        )
        for i in range(n)
    ]


# ── output_csv tests ───────────────────────────────────────────────────────────

class TestFormatSalary:
    def test_min_and_max(self):
        from src.output_csv import format_salary
        job = _make_job(salary_min=800000, salary_max=1200000)
        result = format_salary(job)
        assert "800,000" in result
        assert "1,200,000" in result

    def test_min_only(self):
        from src.output_csv import format_salary
        job = _make_job(salary_min=600000)
        result = format_salary(job)
        assert "600,000" in result
        assert "+" in result

    def test_salary_note(self):
        from src.output_csv import format_salary
        job = _make_job(salary_note="Rs.8L - Rs.12L PA")
        result = format_salary(job)
        assert "Rs.8L" in result

    def test_no_salary(self):
        from src.output_csv import format_salary
        job = _make_job()
        assert format_salary(job) == ""


class TestWriteShortlistCsv:
    def test_creates_dated_and_latest_files(self, tmp_path):
        from src.output_csv import write_shortlist_csv
        jobs = _make_jobs(3, base_score=0.80)
        dated, shortlist = write_shortlist_csv(jobs, min_score=0.65, output_dir=tmp_path)
        assert dated.exists()
        assert (tmp_path / "shortlist_latest.csv").exists()

    def test_filters_below_threshold(self, tmp_path):
        from src.output_csv import write_shortlist_csv
        jobs = _make_jobs(5, base_score=0.80)  # scores: 0.80, 0.75, 0.70, 0.65, 0.60
        _, shortlist = write_shortlist_csv(jobs, min_score=0.70, output_dir=tmp_path)
        # 0.80, 0.75, 0.70 pass; 0.65 and 0.60 do not
        assert len(shortlist) == 3

    def test_all_above_threshold_included(self, tmp_path):
        from src.output_csv import write_shortlist_csv
        jobs = _make_jobs(4, base_score=0.90)  # all >= 0.65
        _, shortlist = write_shortlist_csv(jobs, min_score=0.65, output_dir=tmp_path)
        assert len(shortlist) == 4

    def test_csv_has_correct_columns(self, tmp_path):
        from src.output_csv import write_shortlist_csv, CSV_COLUMNS
        jobs = _make_jobs(2)
        dated, _ = write_shortlist_csv(jobs, min_score=0.65, output_dir=tmp_path)
        with open(dated, encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            headers = reader.fieldnames
        assert headers == CSV_COLUMNS

    def test_csv_row_count_matches_shortlist(self, tmp_path):
        from src.output_csv import write_shortlist_csv
        jobs = _make_jobs(4, base_score=0.80)
        dated, shortlist = write_shortlist_csv(jobs, min_score=0.65, output_dir=tmp_path)
        with open(dated, encoding="utf-8-sig") as f:
            rows = list(csv.DictReader(f))
        assert len(rows) == len(shortlist)

    def test_rank_is_sequential(self, tmp_path):
        from src.output_csv import write_shortlist_csv
        jobs = _make_jobs(3)
        dated, _ = write_shortlist_csv(jobs, min_score=0.65, output_dir=tmp_path)
        with open(dated, encoding="utf-8-sig") as f:
            rows = list(csv.DictReader(f))
        ranks = [int(r["rank"]) for r in rows]
        assert ranks == [1, 2, 3]

    def test_skills_joined_with_comma(self, tmp_path):
        from src.output_csv import write_shortlist_csv
        job = _make_job(matched=["Python", "SQL", "PyTorch"])
        dated, _ = write_shortlist_csv([job], min_score=0.65, output_dir=tmp_path)
        with open(dated, encoding="utf-8-sig") as f:
            row = next(csv.DictReader(f))
        assert row["matched_skills"] == "Python, SQL, PyTorch"

    def test_url_preserved(self, tmp_path):
        from src.output_csv import write_shortlist_csv
        job = _make_job(url="https://jobs.example.com/apply/12345")
        dated, _ = write_shortlist_csv([job], min_score=0.65, output_dir=tmp_path)
        with open(dated, encoding="utf-8-sig") as f:
            row = next(csv.DictReader(f))
        assert row["url"] == "https://jobs.example.com/apply/12345"

    def test_empty_shortlist_still_creates_file(self, tmp_path):
        from src.output_csv import write_shortlist_csv
        jobs = _make_jobs(2, base_score=0.40)  # all below 0.65
        dated, shortlist = write_shortlist_csv(jobs, min_score=0.65, output_dir=tmp_path)
        assert dated.exists()
        assert shortlist == []

    def test_utf8_bom_for_excel_compat(self, tmp_path):
        from src.output_csv import write_shortlist_csv
        job = _make_job(company="Infosys BPM")
        dated, _ = write_shortlist_csv([job], min_score=0.65, output_dir=tmp_path)
        raw = dated.read_bytes()
        # utf-8-sig files start with BOM: EF BB BF
        assert raw[:3] == b"\xef\xbb\xbf"


# ── output_xlsx tests ──────────────────────────────────────────────────────────

class TestWriteShortlistXlsx:
    def test_creates_dated_and_latest_files(self, tmp_path):
        from src.output_xlsx import write_shortlist_xlsx
        jobs = _make_jobs(3)
        dated, _ = write_shortlist_xlsx(jobs, min_score=0.65, output_dir=tmp_path)
        assert dated.exists()
        assert (tmp_path / "shortlist_latest.xlsx").exists()

    def test_xlsx_has_two_sheets(self, tmp_path):
        import openpyxl
        from src.output_xlsx import write_shortlist_xlsx
        jobs = _make_jobs(3)
        dated, _ = write_shortlist_xlsx(jobs, min_score=0.65, output_dir=tmp_path)
        wb = openpyxl.load_workbook(dated)
        assert "Shortlist" in wb.sheetnames
        assert "All Scored" in wb.sheetnames

    def test_shortlist_sheet_row_count(self, tmp_path):
        import openpyxl
        from src.output_xlsx import write_shortlist_xlsx
        jobs = _make_jobs(4, base_score=0.80)  # 0.80, 0.75, 0.70, 0.65 — all pass 0.65
        dated, shortlist = write_shortlist_xlsx(jobs, min_score=0.65, output_dir=tmp_path)
        wb = openpyxl.load_workbook(dated)
        ws = wb["Shortlist"]
        # max_row includes header
        assert ws.max_row == len(shortlist) + 1

    def test_all_scored_sheet_has_all_jobs(self, tmp_path):
        import openpyxl
        from src.output_xlsx import write_shortlist_xlsx
        jobs = _make_jobs(5, base_score=0.80)
        dated, _ = write_shortlist_xlsx(jobs, min_score=0.90, output_dir=tmp_path)
        wb = openpyxl.load_workbook(dated)
        ws = wb["All Scored"]
        # All 5 jobs in this sheet regardless of threshold
        assert ws.max_row == 6  # 1 header + 5 data rows

    def test_header_row_is_frozen(self, tmp_path):
        import openpyxl
        from src.output_xlsx import write_shortlist_xlsx
        jobs = _make_jobs(2)
        dated, _ = write_shortlist_xlsx(jobs, min_score=0.65, output_dir=tmp_path)
        wb = openpyxl.load_workbook(dated)
        ws = wb["Shortlist"]
        assert ws.freeze_panes == "A2"

    def test_url_cell_has_hyperlink(self, tmp_path):
        import openpyxl
        from src.output_xlsx import write_shortlist_xlsx, SHORTLIST_COLS
        job = _make_job(url="https://jobs.example.com/apply/999")
        dated, _ = write_shortlist_xlsx([job], min_score=0.65, output_dir=tmp_path)
        wb = openpyxl.load_workbook(dated)
        ws = wb["Shortlist"]
        # Find URL column index
        url_col = next(i + 1 for i, (_, attr, _) in enumerate(SHORTLIST_COLS) if attr == "url")
        cell = ws.cell(row=2, column=url_col)
        assert cell.hyperlink is not None
        assert "example.com" in str(cell.hyperlink.target)

    def test_fit_score_cell_is_numeric(self, tmp_path):
        import openpyxl
        from src.output_xlsx import write_shortlist_xlsx, SHORTLIST_COLS
        job = _make_job(fit_score=0.847)
        dated, _ = write_shortlist_xlsx([job], min_score=0.65, output_dir=tmp_path)
        wb = openpyxl.load_workbook(dated)
        ws = wb["Shortlist"]
        score_col = next(i + 1 for i, (_, attr, _) in enumerate(SHORTLIST_COLS) if attr == "fit_score")
        cell = ws.cell(row=2, column=score_col)
        assert isinstance(cell.value, float)
        assert cell.value == pytest.approx(0.847)