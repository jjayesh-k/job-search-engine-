"""
output_xlsx.py — Writes the daily shortlist to a formatted Excel workbook.

Features:
  - Colour-coded rows by fit label (Excellent=green, Strong=blue,
    Good=amber, Fair=gray, Weak=light red)
  - Clickable hyperlinks in the URL column (opens job page directly)
  - Frozen header row + column widths auto-fitted to content
  - Separate "Full Scored" sheet with all jobs (above + below threshold)
  - Clean Arial font throughout, professional header style
  - Salary, matched skills, missing skills, LLM summary all included

Output:
  outputs/shortlist_YYYY-MM-DD.xlsx
  outputs/shortlist_latest.xlsx

Requires:
  pip install openpyxl==3.1.4   (add to requirements.txt)
"""

import logging
from datetime import date
from pathlib import Path
from typing import Optional

log = logging.getLogger(__name__)

OUTPUT_DIR = Path("outputs")

# ── Colour palette (openpyxl uses ARGB hex, no leading #) ─────────────────────
# Row fill colours by fit label
LABEL_FILLS = {
    "Excellent": "FFE1F5EE",   # teal tint
    "Strong":    "FFE6F1FB",   # blue tint
    "Good":      "FFFAEEDA",   # amber tint
    "Fair":      "FFF1EFE8",   # gray tint
    "Weak":      "FFFCEBEB",   # red tint
}

# Matching dark text colours per label (for the fit_label cell)
LABEL_FONT_COLORS = {
    "Excellent": "FF0F6E56",   # dark teal
    "Strong":    "FF185FA5",   # dark blue
    "Good":      "FFBA7517",   # dark amber
    "Fair":      "FF5F5E5A",   # dark gray
    "Weak":      "FFA32D2D",   # dark red
}

# Header row
HEADER_FILL   = "FF2C2C2A"   # near-black
HEADER_FONT   = "FFF1EFE8"   # off-white

# Alternating row tint for the full-scored sheet (very subtle)
ALT_ROW_FILL  = "FFF8F7F4"

# ── Column definitions ─────────────────────────────────────────────────────────
# (header label, attribute name on ScoredJob, suggested column width)
SHORTLIST_COLS = [
    ("Rank",           "rank",           6),
    ("Fit Score",      "fit_score",      10),
    ("Fit Label",      "fit_label",      12),
    ("Title",          "title",          34),
    ("Company",        "company",        24),
    ("Location",       "location",       20),
    ("Matched Skills", "matched_skills", 36),
    ("Missing Skills", "missing_skills", 36),
    ("Salary",         "salary",         22),
    ("Posted",         "posted_date",    13),
    ("Source",         "source",         10),
    ("Apply Link",     "url",            14),
    ("AI Summary",     "llm_summary",    50),
]

# Slimmer version for the full-scored sheet (drop LLM summary to save width)
FULL_COLS = [
    ("Rank",           "rank",           6),
    ("Fit Score",      "fit_score",      10),
    ("Fit Label",      "fit_label",      12),
    ("Title",          "title",          34),
    ("Company",        "company",        24),
    ("Location",       "location",       18),
    ("Matched Skills", "matched_skills", 30),
    ("Missing Skills", "missing_skills", 30),
    ("Salary",         "salary",         22),
    ("Posted",         "posted_date",    13),
    ("Source",         "source",         10),
    ("Apply Link",     "url",            14),
]


# ── openpyxl style helpers ─────────────────────────────────────────────────────

def _import_openpyxl():
    """Import openpyxl lazily so the module loads even if not installed."""
    try:
        import openpyxl
        from openpyxl.styles import (
            Font, PatternFill, Alignment, Border, Side
        )
        from openpyxl.utils import get_column_letter
        return openpyxl, Font, PatternFill, Alignment, Border, Side, get_column_letter
    except ImportError:
        raise ImportError(
            "openpyxl is required for Excel output.\n"
            "Install it:  pip install openpyxl==3.1.4\n"
            "Then add to requirements.txt: openpyxl==3.1.4"
        )


def _header_style(Font, PatternFill, Alignment):
    """Return style objects for the header row."""
    font  = Font(name="Arial", bold=True, color=HEADER_FONT, size=10)
    fill  = PatternFill("solid", fgColor=HEADER_FILL)
    align = Alignment(horizontal="center", vertical="center", wrap_text=False)
    return font, fill, align


def _row_fill(label: str, PatternFill):
    """Return a PatternFill for a given fit label."""
    color = LABEL_FILLS.get(label, "FFFFFFFF")
    return PatternFill("solid", fgColor=color)


def _label_font(label: str, Font):
    """Return a bold, coloured Font for the fit_label cell."""
    color = LABEL_FONT_COLORS.get(label, "FF2C2C2A")
    return Font(name="Arial", bold=True, color=color, size=10)


def _thin_border(Border, Side):
    """Return a thin bottom border for header separation."""
    thin = Side(style="thin", color="FFD3D1C7")
    return Border(bottom=thin)


# ── Cell value resolver ────────────────────────────────────────────────────────

def _cell_value(job, attr: str, rank: int):
    """
    Resolve the value for a given column attribute on a ScoredJob.
    Handles special cases: rank injection, list→string, fit_score rounding.
    """
    if attr == "rank":
        return rank
    if attr == "fit_score":
        return round(job.fit_score, 3)
    if attr == "fit_label":
        return job.fit_label()
    if attr == "salary":
        from src.output_csv import format_salary
        return format_salary(job)
    if attr == "matched_skills":
        return ", ".join(job.matched_skills) if job.matched_skills else ""
    if attr == "missing_skills":
        return ", ".join(job.missing_skills) if job.missing_skills else ""
    return getattr(job, attr, "") or ""


# ── Sheet writer ───────────────────────────────────────────────────────────────

def _write_sheet(ws, jobs: list, cols: list, rank_offset: int = 0):
    """
    Write a list of ScoredJob rows onto an openpyxl worksheet.

    Args:
        ws:          The worksheet to write to.
        jobs:        List of ScoredJob objects.
        cols:        Column definitions list (header, attr, width).
        rank_offset: Added to enumerate index (0 for shortlist, 0 for full).
    """
    openpyxl, Font, PatternFill, Alignment, Border, Side, get_column_letter = _import_openpyxl()

    h_font, h_fill, h_align = _header_style(Font, PatternFill, Alignment)
    thin_border = _thin_border(Border, Side)

    # ── Header row ─────────────────────────────────────────────────────────────
    for col_idx, (header, attr, width) in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = h_font
        cell.fill      = h_fill
        cell.alignment = h_align
        cell.border    = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    # Header row height
    ws.row_dimensions[1].height = 22

    # ── Data rows ──────────────────────────────────────────────────────────────
    url_col_idx = next(
        (i + 1 for i, (_, attr, _) in enumerate(cols) if attr == "url"), None
    )

    for row_idx, job in enumerate(jobs, start=2):
        rank  = row_idx - 1 + rank_offset
        label = job.fit_label()
        fill  = _row_fill(label, PatternFill)

        for col_idx, (header, attr, width) in enumerate(cols, start=1):
            value = _cell_value(job, attr, rank)
            cell  = ws.cell(row=row_idx, column=col_idx, value=value)

            # Row background
            cell.fill = fill

            # Base font
            cell.font = Font(name="Arial", size=10, color="FF2C2C2A")

            # Fit label cell gets coloured bold font
            if attr == "fit_label":
                cell.font = _label_font(label, Font)
                cell.alignment = Alignment(horizontal="center", vertical="top")

            # Fit score — right-align, number format
            elif attr == "fit_score":
                cell.number_format = "0.000"
                cell.alignment     = Alignment(horizontal="center", vertical="top")

            # Rank — centre
            elif attr == "rank":
                cell.alignment = Alignment(horizontal="center", vertical="top")

            # URL column — insert hyperlink + display "Apply →"
            elif attr == "url" and value:
                cell.value     = "Apply →"
                cell.hyperlink = str(value)
                cell.font      = Font(
                    name="Arial", size=10,
                    color="FF185FA5",   # blue link colour
                    underline="single",
                )
                cell.alignment = Alignment(horizontal="center", vertical="top")

            # Multi-line text columns — wrap
            elif attr in ("matched_skills", "missing_skills", "llm_summary"):
                cell.alignment = Alignment(
                    horizontal="left", vertical="top", wrap_text=True
                )

            else:
                cell.alignment = Alignment(
                    horizontal="left", vertical="top", wrap_text=False
                )

        # Row height — taller for wrapped columns
        ws.row_dimensions[row_idx].height = 32

    # ── Auto-filter on header ──────────────────────────────────────────────────
    last_col = get_column_letter(len(cols))
    ws.auto_filter.ref = f"A1:{last_col}1"


# ── Main writer ────────────────────────────────────────────────────────────────

def write_shortlist_xlsx(
    scored_jobs: list,
    min_score: Optional[float] = None,
    output_dir: Path = OUTPUT_DIR,
) -> tuple[Path, list]:
    """
    Write the daily shortlist to a formatted Excel workbook.

    Creates two sheets:
      "Shortlist"   — all jobs above min_score, fully formatted
      "All Scored"  — every job scored today (for reference)

    Args:
        scored_jobs: Sorted list of ScoredJob (highest fit first).
        min_score:   Minimum fit_score to include in Shortlist sheet.
        output_dir:  Directory to write into (created if missing).

    Returns:
        (dated_xlsx_path, shortlist) — path written and jobs in the Shortlist sheet.
    """
    openpyxl, *_ = _import_openpyxl()

    if min_score is None:
        from src.config import MIN_FIT_SCORE
        min_score = MIN_FIT_SCORE

    output_dir.mkdir(parents=True, exist_ok=True)

    shortlist = [j for j in scored_jobs if j.fit_score >= min_score]

    if not shortlist:
        log.warning(
            f"No jobs above threshold {min_score} — Excel will be empty. "
            "Lower MIN_FIT_SCORE in .env to see results."
        )

    wb = openpyxl.Workbook()

    # ── Sheet 1: Shortlist ─────────────────────────────────────────────────────
    ws_short = wb.active
    ws_short.title = "Shortlist"
    ws_short.sheet_view.showGridLines = False
    _write_sheet(ws_short, shortlist, SHORTLIST_COLS)

    # ── Sheet 2: All Scored ────────────────────────────────────────────────────
    ws_full = wb.create_sheet("All Scored")
    ws_full.sheet_view.showGridLines = False
    _write_sheet(ws_full, scored_jobs, FULL_COLS)

    # ── Save ───────────────────────────────────────────────────────────────────
    today       = date.today().isoformat()
    dated_path  = output_dir / f"shortlist_{today}.xlsx"
    latest_path = output_dir / "shortlist_latest.xlsx"

    wb.save(dated_path)
    wb.save(latest_path)

    log.info(f"Excel written : {dated_path}  ({len(shortlist)} shortlist rows, {len(scored_jobs)} total)")
    log.info(f"Latest Excel  : {latest_path}")

    return dated_path, shortlist


# ── Standalone test ────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import json
    logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")

    scored_path = Path("outputs/scored_jobs.json")
    if not scored_path.exists():
        print("Run Step 3 first: python main.py --score-only")
        raise SystemExit(1)

    with open(scored_path, encoding="utf-8") as f:
        data = json.load(f)

    from src.score_jobs import ScoredJob
    jobs = [
        ScoredJob(**{k: v for k, v in d.items() if k != "fit_label"})
        for d in data
    ]

    xlsx_path, shortlist = write_shortlist_xlsx(jobs)
    print(f"\nExcel written: {xlsx_path}")
    print(f"Shortlist rows: {len(shortlist)}")
    print(f"Open it: start {xlsx_path}")